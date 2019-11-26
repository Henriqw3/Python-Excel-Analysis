import openpyxl
from datetime import date, datetime

############## Functions to Struct Data #################


def dataStruct(wb, data=False):
    """
    -> Cria a estrutura da área de planilhas, podendo
       retornar apenas a noção das páginas dentro do 
       arquivo ou juntamente com os respectivos dados
    :param wb: A área de trabalho que contém os sheets e os seus dados
    :param datas=False: A extração dos dados não é feita
    :return sheet_names: retorna uma lista com os nomes das páginas do xlsx
    :param datas=True: A extração dos dados é realizada
    :return sheet: dicionário contendo todos os dados da workspace
    @author:Henrique Oliveira
    """

    sheet_names = wb.sheetnames
    
    if data:
        sheet = dict()
        for name in sheet_names:
            sheet[f'{name}'] = dataExtract(wb, name)#Preenche as planilhas
        return sheet
    else:
        return sheet_names


def dataExtract(xlsx, name_sheet):
	"""
	-> Extrai todos os dados das respetivas planilhas
			dentro do arquivo excell
			:param xlsx: a área de trabalho do arquivo (workspace)
			:param name_sheet: a planilha alvo da workspace onde ocorrerá a extração dos dados
			@author:Henrique Oliveira
	"""
	space = wb[f'{name_sheet}']

	data = dict()
	#trunk = 2 .. len(lines)
	#chip = 2 .. len(columns)

	for trunk in range(2,space.max_row+1):#Primeira coluna(column = 1) recebem dicionários
			data[f'{space.cell(row=trunk, column=1).value}'] = dict()
			for chip in range(2, space.max_column+1):#Primeira linha(row=1) recebem dicionários, em seguida os dados
							if name_sheet not in 'Datas':
									data[f'{space.cell(row=trunk, column=1).value}'][
													f'{space.cell(row=1, column=chip).value}'] = space.cell(row=trunk, column=chip).value
							else:
									dte = space.cell(row=trunk, column=chip).value
									dte = ''.join(i for i in dte if i.isdigit() or i ==',')
									dte = dte.replace(',', ' ').split()
									if int(dte[1]) < 10: dte[1] = '0'+dte[1]
									if int(dte[2]) < 10: dte[2] = '0'+dte[2]
									data[f'{space.cell(row=trunk, column=1).value}'][
													f'{space.cell(row=1, column=chip).value}'] = datetime(int(dte[0]), int(dte[1]) , int(dte[2])).date()
	return data


##############    DATA ANALYSIS    ################

def analyseDate (dates):

	"""
	-> Analisa as datas e faz o cálculo de quantos
			dias faltam para a próxima prova existente
	:param dates: dicionário contendo as datas de cada 
	:return dif_days: dicionário contendo o restante de dias 
										para a prova de cada matéria
	@author:Henrique Oliveira
	"""

	atual = datetime(2019, 6, 14).date()
	dif_days = dict()

	for tier in dates:
		first_next = 0
		for v in dates[tier]:
			if first_next > 0:
				break
			if dates[tier][v] >= atual:
				dif_days[tier] = abs((dates[tier][v] - atual).days)
				first_next += 1
	return dif_days



def daysToSub(dates):

	"""
	-> Calcula os dias restantes para a prova substitutiva
	:param dates: recebe as datas em um dicionário para verificação
	:return days_left: retorna um dicionário com os dias restantes
	@author:Henrique Oliveira
	"""

	atual = datetime(2019, 6, 14).date()
	days_left = dict()
	for k, v in dates.items():
		days_left[k] = abs((dates[k]['SUB'] - atual).days)
	
	return days_left

def sumPoints(score):
	"""
	-> Função que soma os pontos para cada matéria
	:param score: dicionário contendo as notas para soma
	:return tot_score: um dicionário com as somas 
	@author:Henrique Oliveira
	"""

	tot_score = dict()

	for tier in score:
		sum_points = 0
		for v in score[tier]:
			if score[tier][v] == None or type(score[tier][v]) == str :
				break
			else:
				sum_points += score[tier][v]
		tot_score[f'{tier}'] = sum_points

	return tot_score


def analyseScore(values_test, score):

	"""
	-> Analisa quantos pontos faltam para aprovação em cada matéria
	:param values_test: dicionário contendo o valor de nota total de cada matéria
	:param score: dicionário contendo o valor de nota que o aluno possui 
	:return analyse: retorna o restante de pontos em um dicionário
	:exception return: caso as prateleiras de cada dicionário não correspondam
	@author:Henrique Oliveira
	"""
	total = sumPoints(values_test)
	try:
		analyse = dict()
		for k, v in score.items():
			if total[k] == 100:
				media = 60
				if v < media:
					analyse[f'{k}'] = media - v
				else : analyse[f'{k}'] = 'nada'
			else:
				media = 60*3
				if v < media:
					analyse[f'{k}'] = media - v
				else : analyse[f'{k}'] = 'nada'
		return analyse
	except:
		return 'Error Values:\n\tDictionaries do not have the same structure\n\t>>Keys must match<<'


def analyseSituation(dates, test_values, points_left):

	"""
	-> Analisa a situação de aprovação do aluno em cada matéria
	:param dates: dicionario contendo as datas
	:param values_test: dicionário contendo a nota total de cada matéria
	:param points_left: dicionário  contendo as notas tiradas pelo aluno 
	:return situation: dicionário contendo a situação final do aluno
	@author:Henrique Oliveira
	"""

	try:

		values_test = sumPoints(test_values)
		tosub = daysToSub(dates)
		situation = dict()
		for k, v in tosub.items():
			if values_test[k] == 100: media = 60
			else: media = 180
			if points_left[k] > media and v < 10:
				situation[k] = 'Reprovado'
			elif 0 < points_left[k] < media and v < 10:
				situation[k] = 'Recuperação'
			else:
				situation[k] = 'Aprovado'
		
		return situation 
	except:
		return 'Error Values:\n\tDictionaries do not have the same structure\n\t>>Keys must match<<'

##############   ANALYSIS REPORT   ################

def passFile(dates,pts_left,situation, values_test):

	"""
	-> Passa os dados processados para o relatório do aluno
	"""

	report = open('relatorio.txt', 'w')

	report.write  ('*+*+*+*+*+*+*+* STUDENT REPORT *+*+*+*+*+*+*+*\n')
	for k, v in situation.items():
		report.write(f'\n>> {k} : \n')
		if pts_left[k] > 0 and situation[k] not in 'Reprovado':
			report.write(f'\n\tTotal de {pts_left[k]} pontos necessários para passar.\n')
			report.write(f'\tTotal de {dates[k]} dias para a próxima prova.\n')
			if values_test[k]['SUB'] == 'Não tem sub':
				report.write('\n\tNão haverá prova SUB!\n')
		elif situation[k] == 'Reprovado':
			report.write(f'\n\tSituação: Reprovado nesta matéria.\n')

def finalCommands(dataSheet):
	next_test = analyseDate(dataSheet['Datas'])
	notas = sumPoints(dataSheet['Notas'])
	pts_restantes = analyseScore(dataSheet['Valores'], notas)
	situacao = analyseSituation(dataSheet['Datas'],dataSheet['Valores'] , pts_restantes)

	passFile(next_test,pts_restantes,situacao, dataSheet['Valores'])	

############## ******* END ********* ################

# ~/ main program

wb = openpyxl.load_workbook('provas.xlsx')#Arquivo xlsx

dataSheet = dataStruct(wb, data = True)#Dados extraídos

finalCommands(dataSheet)
